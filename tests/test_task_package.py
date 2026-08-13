from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path
from unittest import mock

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from dingmail.task_models import MailTask
from dingmail.task_package import (
    TASKS_FILENAME,
    TASKS_SHEET_NAME,
    ensure_unique_task_ids,
    load_tasks_from_package,
    package_relpath,
    resolve_user_path,
    save_tasks_to_package,
)


class TaskPackageTests(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory(prefix="dingmail_task_package_")
        self.addCleanup(self._tmp.cleanup)
        self.package_dir = Path(self._tmp.name)
        (self.package_dir / "content").mkdir()
        (self.package_dir / "attachments").mkdir()
        (self.package_dir / "content" / "body.md").write_text("# 标题\n\n正文", encoding="utf-8")

    def test_resolve_user_path_rejects_paths_outside_package(self) -> None:
        outside = self.package_dir.parent / "outside.md"
        outside.write_text("x", encoding="utf-8")

        with self.assertRaisesRegex(ValueError, "任务包目录内"):
            resolve_user_path(self.package_dir, str(outside))

        with self.assertRaisesRegex(ValueError, "任务包目录内"):
            resolve_user_path(self.package_dir, "..\\outside.md")

    def test_package_relpath_rejects_paths_outside_package(self) -> None:
        outside = self.package_dir.parent / "outside.md"
        outside.write_text("x", encoding="utf-8")

        with self.assertRaisesRegex(ValueError, "任务包目录内"):
            package_relpath(self.package_dir, outside)

    def test_load_tasks_prefers_named_tasks_sheet(self) -> None:
        workbook = openpyxl.Workbook()
        meta = workbook.active
        meta.title = "Meta"
        meta["A1"] = "not tasks"

        tasks = workbook.create_sheet(TASKS_SHEET_NAME)
        tasks.append(
            ["任务ID", "是否启用", "收件人", "抄送人", "主题", "开头/补充内容", "Markdown路径", "是否有附件", "附件路径", "是否定时发送", "定时发送时间", "备注"]
        )
        tasks.append(["task-1", "是", "a@example.com", "", "主题", "", "content/body.md", "否", "", "否", "", ""])
        workbook.active = 0
        workbook.save(self.package_dir / TASKS_FILENAME)
        workbook.close()

        loaded = load_tasks_from_package(self.package_dir)
        self.assertEqual(1, len(loaded))
        self.assertEqual("task-1", loaded[0].task_id)
        self.assertEqual(["a@example.com"], loaded[0].to_recipients)

    def test_save_tasks_preserves_other_sheets(self) -> None:
        workbook = openpyxl.Workbook()
        tasks = workbook.active
        tasks.title = TASKS_SHEET_NAME
        tasks["A1"] = "old"
        tasks.freeze_panes = "B2"
        tasks.column_dimensions["A"].width = 24
        meta = workbook.create_sheet("Meta")
        meta["A1"] = "keep me"
        workbook.save(self.package_dir / TASKS_FILENAME)
        workbook.close()

        save_tasks_to_package(
            self.package_dir,
            [
                MailTask(
                    task_id="task-1",
                    to_recipients=["a@example.com"],
                    subject="主题",
                    markdown_path="content/body.md",
                )
            ],
        )

        saved = openpyxl.load_workbook(self.package_dir / TASKS_FILENAME)
        self.addCleanup(saved.close)
        self.assertIn(TASKS_SHEET_NAME, saved.sheetnames)
        self.assertIn("Meta", saved.sheetnames)
        self.assertEqual("keep me", saved["Meta"]["A1"].value)
        self.assertEqual("任务ID", saved[TASKS_SHEET_NAME]["A1"].value)
        self.assertEqual("task-1", saved[TASKS_SHEET_NAME]["A2"].value)
        self.assertEqual("B2", saved[TASKS_SHEET_NAME].freeze_panes)
        self.assertEqual(24, saved[TASKS_SHEET_NAME].column_dimensions["A"].width)

    def test_save_tasks_preserves_unknown_task_sheet_columns_by_task_id(self) -> None:
        workbook = openpyxl.Workbook()
        tasks = workbook.active
        tasks.title = TASKS_SHEET_NAME
        tasks.append(
            ["任务ID", "是否启用", "收件人", "抄送人", "主题", "开头/补充内容", "Markdown路径", "是否有附件", "附件路径", "是否定时发送", "定时发送时间", "备注", "人工复核备注"]
        )
        tasks.append(["task-1", "是", "old@example.com", "", "旧主题", "", "content/body.md", "否", "", "否", "", "", "保留这列"])
        workbook.save(self.package_dir / TASKS_FILENAME)
        workbook.close()

        save_tasks_to_package(
            self.package_dir,
            [
                MailTask(
                    task_id="task-1",
                    to_recipients=["a@example.com"],
                    subject="新主题",
                    markdown_path="content/body.md",
                )
            ],
        )

        saved = openpyxl.load_workbook(self.package_dir / TASKS_FILENAME)
        self.addCleanup(saved.close)
        sheet = saved[TASKS_SHEET_NAME]
        # "最近结果" is now standard column 13; user extra "人工复核备注" moves to 14
        self.assertEqual("最近结果", sheet.cell(row=1, column=13).value)
        self.assertEqual("人工复核备注", sheet.cell(row=1, column=14).value)
        self.assertEqual("保留这列", sheet.cell(row=2, column=14).value)
        self.assertEqual("新主题", sheet.cell(row=2, column=5).value)

    def test_ensure_unique_task_ids_repairs_missing_and_duplicate_ids(self) -> None:
        tasks = [
            MailTask(task_id="task-1", markdown_path="content/body.md"),
            MailTask(task_id="task-1", markdown_path="content/body.md"),
            MailTask(task_id="", markdown_path="content/body.md"),
        ]

        repairs = ensure_unique_task_ids(tasks)
        self.assertEqual(2, len(repairs))
        self.assertEqual(3, len({task.task_id for task in tasks}))
        self.assertTrue(all(task.task_id for task in tasks))

    def test_save_tasks_repairs_duplicate_and_missing_task_ids(self) -> None:
        tasks = [
            MailTask(task_id="task-1", to_recipients=["a@example.com"], subject="主题1", markdown_path="content/body.md"),
            MailTask(task_id="task-1", to_recipients=["b@example.com"], subject="主题2", markdown_path="content/body.md"),
            MailTask(task_id="", to_recipients=["c@example.com"], subject="主题3", markdown_path="content/body.md"),
        ]

        save_tasks_to_package(self.package_dir, tasks)

        self.assertEqual(3, len({task.task_id for task in tasks}))
        self.assertTrue(all(task.task_id for task in tasks))

        loaded = load_tasks_from_package(self.package_dir)
        self.assertEqual(3, len(loaded))
        self.assertEqual(3, len({task.task_id for task in loaded}))
        self.assertTrue(all(task.task_id for task in loaded))

    def _write_sheet_with_extra_column(self, rows: list[list[object]]) -> None:
        headers = [
            "任务ID", "是否启用", "收件人", "抄送人", "主题", "开头/补充内容",
            "Markdown路径", "是否有附件", "附件路径", "是否定时发送", "定时发送时间", "备注", "部门",
        ]
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = TASKS_SHEET_NAME
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        workbook.save(self.package_dir / TASKS_FILENAME)
        workbook.close()

    def _read_extra_column(self) -> list[tuple[object, object]]:
        workbook = openpyxl.load_workbook(self.package_dir / TASKS_FILENAME)
        self.addCleanup(workbook.close)
        sheet = workbook[TASKS_SHEET_NAME]
        # User extra column "部门" is now at column 14 (after new standard "最近结果" at 13)
        return [
            (sheet.cell(row=row, column=1).value, sheet.cell(row=row, column=14).value)
            for row in range(2, sheet.max_row + 1)
        ]

    def test_load_tasks_keeps_missing_task_ids_empty_for_repair_reporting(self) -> None:
        self._write_sheet_with_extra_column(
            [["", "是", "a@example.com", "", "主题1", "", "content/body.md", "否", "", "否", "", "", "市场部"]]
        )

        loaded = load_tasks_from_package(self.package_dir)
        self.assertEqual("", loaded[0].task_id)

        repairs = ensure_unique_task_ids(loaded)
        self.assertEqual(1, len(repairs))

    def test_save_preserves_extra_columns_when_task_ids_missing(self) -> None:
        self._write_sheet_with_extra_column(
            [
                ["", "是", "a@example.com", "", "主题1", "", "content/body.md", "否", "", "否", "", "", "市场部"],
                ["", "是", "b@example.com", "", "主题2", "", "content/body.md", "否", "", "否", "", "", "研发部"],
            ]
        )

        tasks = load_tasks_from_package(self.package_dir)
        ensure_unique_task_ids(tasks)
        save_tasks_to_package(self.package_dir, tasks)

        extra = self._read_extra_column()
        self.assertEqual(["市场部", "研发部"], [value for _task_id, value in extra])
        self.assertTrue(all(task_id for task_id, _value in extra))

    def test_save_preserves_extra_columns_when_task_ids_duplicated(self) -> None:
        self._write_sheet_with_extra_column(
            [
                ["t-1", "是", "a@example.com", "", "主题1", "", "content/body.md", "否", "", "否", "", "", "市场部"],
                ["t-1", "是", "b@example.com", "", "主题2", "", "content/body.md", "否", "", "否", "", "", "研发部"],
            ]
        )

        tasks = load_tasks_from_package(self.package_dir)
        repairs = ensure_unique_task_ids(tasks)
        self.assertEqual(1, len(repairs))
        save_tasks_to_package(self.package_dir, tasks)

        extra = self._read_extra_column()
        self.assertEqual(("t-1", "市场部"), extra[0])
        self.assertEqual("研发部", extra[1][1])
        self.assertNotEqual("t-1", extra[1][0])

    def test_failed_save_keeps_existing_tasks_file_intact(self) -> None:
        save_tasks_to_package(
            self.package_dir,
            [MailTask(task_id="task-1", to_recipients=["a@example.com"], subject="主题", markdown_path="content/body.md")],
        )
        tasks_path = self.package_dir / TASKS_FILENAME
        before = tasks_path.read_bytes()

        with mock.patch.object(openpyxl.workbook.workbook.Workbook, "save", side_effect=OSError("disk full")):
            with self.assertRaises(OSError):
                save_tasks_to_package(
                    self.package_dir,
                    [MailTask(task_id="task-2", to_recipients=["b@example.com"], subject="主题2", markdown_path="content/body.md")],
                )

        self.assertEqual(before, tasks_path.read_bytes())
        self.assertEqual([], list(self.package_dir.glob("*.tmp-*")))


if __name__ == "__main__":
    unittest.main()
