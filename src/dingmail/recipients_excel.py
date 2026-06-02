from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

import openpyxl

from .model import RecipientsConfig

_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


@dataclass(frozen=True)
class Recipient:
    email: str
    variables: dict[str, str]
    row_number: int


def _header_to_index(ws, header_row: int) -> dict[str, int]:
    headers = []
    for cell in ws[header_row]:
        headers.append("" if cell.value is None else str(cell.value).strip())

    header_to_index: dict[str, int] = {}
    for idx, h in enumerate(headers):
        if h and h not in header_to_index:
            header_to_index[h] = idx
    return header_to_index


def _key_to_index(columns: dict[str, str], header_to_index: dict[str, int]) -> dict[str, int]:
    key_to_index: dict[str, int] = {}
    for key, header_name in columns.items():
        header_name = str(header_name).strip()
        if not header_name:
            continue
        if header_name not in header_to_index:
            raise ValueError(f"Excel 表头缺少列：{header_name!r}（用于字段 {key!r}）")
        key_to_index[key] = header_to_index[header_name]

    if "email" not in key_to_index:
        raise ValueError("recipients.columns 必须包含 email 映射")
    return key_to_index


def _variables_from_row(row: tuple[object, ...], key_to_index: dict[str, int]) -> dict[str, str]:
    variables: dict[str, str] = {}
    for key, col_idx in key_to_index.items():
        value = row[col_idx] if col_idx < len(row) else None
        variables[key] = "" if value is None else str(value).strip()
    return variables


def _recipient_from_row(row: tuple[object, ...], row_idx: int, key_to_index: dict[str, int]) -> Recipient:
    variables = _variables_from_row(row, key_to_index)
    email = variables.get("email", "").strip()
    if not email:
        raise ValueError(f"第 {row_idx} 行 email 为空")
    if not _EMAIL_RE.match(email):
        raise ValueError(f"第 {row_idx} 行 email 非法：{email!r}")
    return Recipient(email=email, variables=variables, row_number=row_idx)


def _is_blank_row(row: tuple[object, ...]) -> bool:
    return not row or all(value is None or str(value).strip() == "" for value in row)


def load_recipients(campaign_dir: Path, cfg: RecipientsConfig) -> list[Recipient]:
    xlsx_path = (campaign_dir / cfg.file).resolve()
    if not xlsx_path.is_file():
        raise FileNotFoundError(f"未找到收件人 Excel：{xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[cfg.sheet] if cfg.sheet else wb.active
    key_to_index = _key_to_index(cfg.columns, _header_to_index(ws, cfg.header_row))
    recipients: list[Recipient] = []
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=cfg.header_row + 1, values_only=True),
        start=cfg.header_row + 1,
    ):
        if _is_blank_row(row):
            continue
        recipients.append(_recipient_from_row(row, row_idx, key_to_index))

    if not recipients:
        raise ValueError("收件人清单为空")

    return recipients
