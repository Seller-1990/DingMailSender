from __future__ import annotations

import os
import uuid
from datetime import datetime
from pathlib import Path

import openpyxl

from .task_models import MailTask, PackageLayout

TASKS_FILENAME = "tasks.xlsx"
TASKS_SHEET_NAME = "Tasks"
PACKAGE_README_FILENAME = "README_操作说明.md"
CONTENT_DIRNAME = "content"
ASSETS_DIRNAME = "assets"
ATTACHMENTS_DIRNAME = "attachments"
DATETIME_FMT = "%Y-%m-%d %H:%M:%S"

TASK_COLUMNS = [
    "任务ID",
    "是否启用",
    "收件人",
    "抄送人",
    "主题",
    "开头/补充内容",
    "Markdown路径",
    "是否有附件",
    "附件路径",
    "是否定时发送",
    "定时发送时间",
    "备注",
    "最近结果",
]

ExtraColumnState = tuple[range, dict[int, object]]


def default_package_layout(package_dir: Path) -> PackageLayout:
    return PackageLayout(
        package_dir=str(package_dir),
        tasks_file=str(package_dir / TASKS_FILENAME),
        content_dir=str(package_dir / CONTENT_DIRNAME),
        assets_dir=str(package_dir / ASSETS_DIRNAME),
        attachments_dir=str(package_dir / ATTACHMENTS_DIRNAME),
        readme_file=str(package_dir / PACKAGE_README_FILENAME),
    )


def split_emails(value: str) -> list[str]:
    raw = (value or "").replace("；", ";").replace(",", ";")
    return [x.strip() for x in raw.split(";") if x.strip()]


def join_emails(values: list[str]) -> str:
    return "; ".join([x.strip() for x in values if str(x).strip()])


def split_paths(value: str) -> list[str]:
    raw = (value or "").replace("\r\n", "\n").replace("；", ";")
    normalized = raw.replace("\n", ";")
    return [x.strip() for x in normalized.split(";") if x.strip()]


def join_paths(values: list[str]) -> str:
    return "; ".join([x.strip() for x in values if str(x).strip()])


def parse_bool(value: object) -> bool:
    if isinstance(value, bool):
        return value
    text = str(value or "").strip().lower()
    return text in {"1", "true", "yes", "y", "是", "启用", "有"}


def parse_datetime(value: object) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.replace(microsecond=0)
    text = str(value).strip()
    if not text:
        return None
    for fmt in (DATETIME_FMT, "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    raise ValueError(f"无法识别的定时发送时间：{text}")


def datetime_to_excel_text(value: datetime | None) -> str:
    if value is None:
        return ""
    return value.strftime(DATETIME_FMT)


def package_relpath(package_dir: Path, path: Path) -> str:
    try:
        return str(path.resolve().relative_to(package_dir.resolve()))
    except ValueError:
        raise ValueError(f"路径必须位于任务包目录内：{path.resolve()}")


_RESOLVED_PACKAGE_ROOTS: dict[Path, Path] = {}


def _package_root(package_dir: Path) -> Path:
    """包根 realpath 按 Path 对象缓存：同一包内成百上千次路径校验共享一次 realpath。"""
    key = Path(package_dir)
    cached = _RESOLVED_PACKAGE_ROOTS.get(key)
    if cached is None:
        cached = key.resolve()
        _RESOLVED_PACKAGE_ROOTS[key] = cached
    return cached


def _resolve_within_package(package_dir: Path, path: Path) -> Path:
    package_root = _package_root(package_dir)
    resolved = path.resolve() if path.is_absolute() else (package_root / path).resolve()
    try:
        resolved.relative_to(package_root)
    except ValueError as exc:
        raise ValueError(f"路径必须位于任务包目录内：{resolved}") from exc
    return resolved


def resolve_user_path(package_dir: Path, raw_path: str) -> Path:
    value = str(raw_path or "").strip()
    if not value:
        raise ValueError("路径为空")
    return _resolve_within_package(package_dir, Path(value))


def _select_tasks_sheet(workbook: openpyxl.Workbook):
    return workbook[TASKS_SHEET_NAME] if TASKS_SHEET_NAME in workbook.sheetnames else workbook.active


def _ensure_tasks_sheet(workbook: openpyxl.Workbook):
    if TASKS_SHEET_NAME in workbook.sheetnames:
        return workbook[TASKS_SHEET_NAME]

    if len(workbook.sheetnames) == 1 and workbook.active.max_row == 1 and workbook.active.max_column == 1:
        only_sheet = workbook.active
        if only_sheet["A1"].value in (None, ""):
            only_sheet.title = TASKS_SHEET_NAME
            return only_sheet

    return workbook.create_sheet(TASKS_SHEET_NAME, index=0)


def ensure_unique_task_ids(tasks: list[MailTask]) -> list[str]:
    seen_ids: dict[str, int] = {}
    repairs: list[str] = []
    for index, task in enumerate(tasks, start=2):
        raw_task_id = str(task.task_id or "").strip()
        if not raw_task_id:
            task.task_id = uuid.uuid4().hex
            repairs.append(f"第 {index} 行缺少任务ID，已自动重置。")
            seen_ids[task.task_id] = index
            continue

        previous_row = seen_ids.get(raw_task_id)
        if previous_row is not None:
            task.task_id = uuid.uuid4().hex
            repairs.append(f"第 {index} 行任务ID与第 {previous_row} 行重复，已自动重置。")
            seen_ids[task.task_id] = index
            continue

        seen_ids[raw_task_id] = index

    return repairs


def _task_row_values(task: MailTask) -> list[object]:
    return [
        task.task_id,
        "是" if task.enabled else "否",
        join_emails(task.to_recipients),
        join_emails(task.cc_recipients),
        task.subject,
        task.intro_text,
        task.markdown_path,
        "是" if task.attachment_count() > 0 else "否",
        join_paths(task.attachment_paths),
        "是" if task.schedule_enabled else "否",
        datetime_to_excel_text(task.scheduled_at),
        task.note,
        task.last_delivery_status,
    ]


def _snapshot_extra_task_columns(
    worksheet,
) -> tuple[range, dict[int, object], dict[str, dict[int, object]], dict[int, dict[int, object]]]:
    """Snapshot user-added columns that are not in TASK_COLUMNS.

    Handles backward compatibility: old files may have fewer standard columns,
    so we identify extra columns by header name rather than position.
    """
    task_columns_set = set(TASK_COLUMNS)
    new_extra_start = len(TASK_COLUMNS) + 1

    # Identify which file columns have headers NOT in our standard set
    file_extra_cols: list[int] = []
    for col in range(1, worksheet.max_column + 1):
        header = worksheet.cell(row=1, column=col).value
        header_str = str(header).strip() if header is not None else ""
        if header_str and header_str not in task_columns_set:
            file_extra_cols.append(col)

    if not file_extra_cols:
        empty_range = range(new_extra_start, new_extra_start)
        return empty_range, {}, {}, {}

    # Remap file extra columns to sequential positions starting after TASK_COLUMNS
    remap: dict[int, int] = {}  # old_col -> new_col
    for offset, old_col in enumerate(file_extra_cols):
        remap[old_col] = new_extra_start + offset

    extra_columns = range(new_extra_start, new_extra_start + len(file_extra_cols))
    extra_headers = {remap[col]: worksheet.cell(row=1, column=col).value for col in file_extra_cols}

    extra_values_by_task_id: dict[str, dict[int, object]] = {}
    extra_values_by_row: dict[int, dict[int, object]] = {}
    for row_index in range(2, worksheet.max_row + 1):
        values = {remap[col]: worksheet.cell(row=row_index, column=col).value for col in file_extra_cols}
        task_id = str(worksheet.cell(row=row_index, column=1).value or "").strip()
        if task_id and task_id not in extra_values_by_task_id:
            extra_values_by_task_id[task_id] = values
        else:
            extra_values_by_row[row_index] = values
    return extra_columns, extra_headers, extra_values_by_task_id, extra_values_by_row


def _resize_task_sheet_rows(worksheet, target_row_count: int) -> None:
    current_row_count = worksheet.max_row
    if current_row_count < target_row_count:
        worksheet.insert_rows(current_row_count + 1, target_row_count - current_row_count)
    elif current_row_count > target_row_count:
        worksheet.delete_rows(target_row_count + 1, current_row_count - target_row_count)


def _write_task_header(worksheet, extra_headers: dict[int, object]) -> None:
    for col_index, value in enumerate(TASK_COLUMNS, start=1):
        worksheet.cell(row=1, column=col_index).value = value
    for col_index, value in extra_headers.items():
        worksheet.cell(row=1, column=col_index).value = value


def _write_task_row(
    worksheet,
    row_index: int,
    task: MailTask,
    extra_state: ExtraColumnState,
) -> None:
    extra_columns, extra_values = extra_state
    for col_index, value in enumerate(_task_row_values(task), start=1):
        worksheet.cell(row=row_index, column=col_index).value = value
    for col_index in extra_columns:
        worksheet.cell(row=row_index, column=col_index).value = extra_values.get(col_index)


def _write_tasks_sheet(worksheet, tasks: list[MailTask]) -> None:
    extra_columns, extra_headers, extra_values_by_task_id, extra_values_by_row = _snapshot_extra_task_columns(
        worksheet
    )
    _resize_task_sheet_rows(worksheet, len(tasks) + 1)
    _write_task_header(worksheet, extra_headers)
    for row_index, task in enumerate(tasks, start=2):
        extra_values = extra_values_by_task_id.get(task.task_id)
        if extra_values is None:
            # 刚被修复过 ID 的行在旧文件里没有可用键；导入修复随即写回时行序未变，按位置取回。
            extra_values = extra_values_by_row.get(row_index, {})
        _write_task_row(
            worksheet,
            row_index=row_index,
            task=task,
            extra_state=(extra_columns, extra_values),
        )


def _read_task_rows(tasks_path: Path) -> list[tuple[object, ...]]:
    if not tasks_path.is_file():
        raise FileNotFoundError(f"未找到任务表：{tasks_path}")

    workbook = openpyxl.load_workbook(tasks_path, data_only=True, read_only=True)
    try:
        worksheet = _select_tasks_sheet(workbook)
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    return rows


def _task_header_map(header_row: tuple[object, ...]) -> dict[str, int]:
    headers = [str(value).strip() if value is not None else "" for value in header_row]
    return {name: index for index, name in enumerate(headers) if name}


def _row_value(row: tuple[object, ...], header_map: dict[str, int], column: str) -> object:
    index = header_map.get(column)
    if index is None or index >= len(row):
        return None
    return row[index]


def _is_blank_task_row(row: tuple[object, ...]) -> bool:
    return not row or all(value is None or str(value).strip() == "" for value in row)


def _mail_task_from_row(row: tuple[object, ...], header_map: dict[str, int]) -> MailTask:
    def value(column: str) -> object:
        return _row_value(row, header_map, column)

    scheduled_at: datetime | None = None
    schedule_raw = str(value("定时发送时间") or "").strip()
    try:
        scheduled_at = parse_datetime(schedule_raw)
    except ValueError:
        # 无法解析的定时时间不能静默清空（会在保存时覆盖原值丢数据）：
        # 保留原始文本到备注列，提醒用户手动修正
        return MailTask(
            task_id=str(value("任务ID") or "").strip(),
            enabled=parse_bool(value("是否启用")) if value("是否启用") is not None else True,
            to_recipients=split_emails(str(value("收件人") or "")),
            cc_recipients=split_emails(str(value("抄送人") or "")),
            subject=str(value("主题") or "").strip(),
            intro_text=str(value("开头/补充内容") or ""),
            markdown_path=str(value("Markdown路径") or "").strip(),
            attachment_paths=split_paths(str(value("附件路径") or "")),
            schedule_enabled=parse_bool(value("是否定时发送")),
            scheduled_at=None,
            note=(str(value("备注") or "").strip() + f"；[定时发送时间无法解析：{schedule_raw}]").strip("；"),
            last_delivery_status=str(value("最近结果") or "").strip(),
        )

    return MailTask(
        task_id=str(value("任务ID") or "").strip(),
        enabled=parse_bool(value("是否启用")) if value("是否启用") is not None else True,
        to_recipients=split_emails(str(value("收件人") or "")),
        cc_recipients=split_emails(str(value("抄送人") or "")),
        subject=str(value("主题") or "").strip(),
        intro_text=str(value("开头/补充内容") or ""),
        markdown_path=str(value("Markdown路径") or "").strip(),
        attachment_paths=split_paths(str(value("附件路径") or "")),
        schedule_enabled=parse_bool(value("是否定时发送")),
        scheduled_at=scheduled_at,
        note=str(value("备注") or "").strip(),
        last_delivery_status=str(value("最近结果") or "").strip(),
    )


def load_tasks_from_package(package_dir: Path) -> list[MailTask]:
    tasks_path = package_dir / TASKS_FILENAME
    rows = _read_task_rows(tasks_path)
    if not rows:
        return []

    header_map = _task_header_map(rows[0])
    tasks: list[MailTask] = []
    for row in rows[1:]:
        if _is_blank_task_row(row):
            continue
        tasks.append(_mail_task_from_row(row, header_map))

    return tasks


def save_tasks_to_package(package_dir: Path, tasks: list[MailTask]) -> Path:
    package_dir.mkdir(parents=True, exist_ok=True)
    ensure_unique_task_ids(tasks)
    tasks_path = package_dir / TASKS_FILENAME
    workbook = openpyxl.load_workbook(tasks_path) if tasks_path.exists() else openpyxl.Workbook()
    try:
        worksheet = _ensure_tasks_sheet(workbook)
        _write_tasks_sheet(worksheet, tasks)
        # 先写同目录临时文件再原子替换，保存中途失败（断电/被杀/磁盘满）不会截断 tasks.xlsx。
        tmp_path = tasks_path.with_name(f"{TASKS_FILENAME}.tmp-{os.getpid()}")
        try:
            workbook.save(tmp_path)
            os.replace(tmp_path, tasks_path)
        except PermissionError:
            # 文件被 Excel 等程序占用时 os.replace 失败——保留 tmp 让用户可恢复
            raise PermissionError(
                f"无法保存 {tasks_path.name}：文件可能被 Excel 或其他程序占用。\n"
                f"修改内容已暂存到：{tmp_path.name}，请关闭占用程序后重试。"
            ) from None
        except BaseException:
            tmp_path.unlink(missing_ok=True)
            raise
    finally:
        workbook.close()
    return tasks_path
